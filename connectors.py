# -*- coding: utf-8 -*-
"""Сбор данных. Схема fallback по ТЗ:
официальный API -> официальный XLSX/CSV -> автоматизированный браузер -> ручной ввод.
Здесь реализованы: YouTube API (если задан ключ), CSV-импорт, ручной ввод.
Остальные источники (Instagram Graph API, Лайфдюн, GetCourse) подключаются
через настройки токенов тем же интерфейсом коллектора."""
import csv, io, requests
from datetime import date, datetime
from db import db, MetricSnapshot, RunLog, Notification, get_setting, Channel

DAILY_METRICS = ["followers", "views", "impressions", "reach", "opens", "reads",
                 "likes", "comments", "saves", "shares", "reactions",
                 "subscribed", "unsubscribed"]

# Применимые метрики по платформам: отсутствующая применимая -> MISSING,
# неприменимая -> NOT_AVAILABLE (не является проблемой сбора).
BASE = {"followers", "views", "reach", "likes", "comments", "shares", "subscribed", "unsubscribed"}
PLATFORM_METRICS = {
    "instagram": BASE | {"impressions", "saves", "opens"},
    "youtube": BASE | {"impressions"},
    "max": BASE | {"opens"},
    "telegram": BASE | {"opens", "reads"},
    "tiktok": BASE | {"saves"},
    "vk": BASE | {"impressions", "saves", "opens"},
    "dzen": BASE | {"impressions", "opens", "reads"},
}


def start_run(kind):
    r = RunLog(kind=kind, status="OK")
    db.session.add(r)
    db.session.commit()
    return r.id


def finish_run(run_id, details="", status="OK"):
    r = RunLog.query.get(run_id)
    r.status = status
    r.details = details
    db.session.commit()


def save_metric(run_id, channel_id, d, metric, value, source, status="OK"):
    db.session.add(MetricSnapshot(channel_id=channel_id, date=d, metric=metric,
                                  value=value, status=status, source=source, run_id=run_id))


def collect_youtube(run_id):
    """YouTube Data API: подписчики + дневная дельта просмотров (viewCount кумулятивный).
    Канал резолвится по хэндлу @AlexeySyrover, id кэшируется в настройках."""
    key = get_setting("youtube_api_key")
    ch = Channel.query.filter(Channel.platform == "youtube", Channel.is_active == True).first()  # noqa
    if not key or not ch:
        return "youtube: ключ API не задан или канал не найден — пропуск"
    try:
        params = {"part": "statistics,snippet", "key": key}
        cid = get_setting("youtube_channel_id", "")
        if cid:
            params["id"] = cid
        else:
            params["forHandle"] = "AlexeySyrover"
        r = requests.get("https://www.googleapis.com/youtube/v3/channels", params=params, timeout=20)
        items = r.json().get("items", [])
        if not items:
            return "youtube: канал не найден"
        it = items[0]
        if not cid:
            from db import set_setting
            set_setting("youtube_channel_id", it["id"])
        st = it.get("statistics", {})
        today = date.today()
        subs = int(st.get("subscriberCount", 0))
        views_total = int(st.get("viewCount", 0))
        save_metric(run_id, ch.id, today, "followers", subs, "youtube_api")
        # дневная дельта просмотров: вычитаем последнюю кумулятивную отметку
        last = (MetricSnapshot.query.filter_by(channel_id=ch.id, metric="yt_views_total")
                .order_by(MetricSnapshot.date.desc(), MetricSnapshot.fetched_at.desc()).first())
        base = last.value if (last and last.value) else None
        if base is None or views_total >= base:
            save_metric(run_id, ch.id, today, "views", max(views_total - (base or views_total), 0),
                        "youtube_api")
        save_metric(run_id, ch.id, today, "yt_views_total", views_total, "youtube_api")
        db.session.commit()
        return f"youtube: OK (подписчиков {subs}, просмотров всего {views_total})"
    except Exception as e:
        db.session.add(Notification(level="error", message=f"Ошибка YouTube API: {e}"))
        db.session.commit()
        return f"youtube: ERROR {e}"


def mark_missing(run_id, d=None):
    """Метим метрики всех каналов за дату как MISSING, если данных нет —
    чтобы отличать '0' от 'не получено'. Итог — сводное уведомление по ТЗ."""
    d = d or date.today()
    missing_by_channel = {}
    no_source = []
    from datetime import date as _date, timedelta as _td
    since = _date.today() - _td(days=60)
    for ch in Channel.query.filter_by(is_active=True).all():
        # итоговый статус каждой метрики = последний снапшот по (канал, дата, метрика)
        latest = {}
        for r in MetricSnapshot.query.filter_by(channel_id=ch.id, date=d).all():
            k = r.metric
            if k not in latest or (r.fetched_at or 0) >= (latest[k].fetched_at or 0):
                latest[k] = r
        # адаптивная применимость: метрика применима, если приходила по каналу
        # хоть раз за 60 дней (реальные возможности источников)
        ever = {r.metric for r in db.session.query(MetricSnapshot.metric).filter(
            MetricSnapshot.channel_id == ch.id, MetricSnapshot.date >= since,
            MetricSnapshot.source.notin_(["demo", "collector"]),
            MetricSnapshot.value.isnot(None)).distinct()}
        base = PLATFORM_METRICS.get(ch.platform, set(DAILY_METRICS)) & set(DAILY_METRICS)
        applicable = (ever & set(DAILY_METRICS)) or base
        ok_today = {m for m, r in latest.items() if r.status == "OK"}
        if ok_today:
            # источник за этот день отчитался: чего не принёс — недоступно, а не потеряно
            miss = []
            for m in DAILY_METRICS:
                if m in ok_today:
                    continue
                cur = latest.get(m)
                if cur is not None and cur.status == "NOT_AVAILABLE":
                    continue
                save_metric(run_id, ch.id, d, m, None, "collector", status="NOT_AVAILABLE")
        elif ever:
            # источник обычно отчитывается, но за эту дату промолчал — честный MISSING
            miss = [m for m in sorted(applicable)
                    if latest.get(m) is None or latest[m].status != "OK"]
            for m in miss:
                cur = latest.get(m)
                if cur is not None and cur.status == "MISSING":
                    continue
                save_metric(run_id, ch.id, d, m, None, "collector", status="MISSING")
            for m in DAILY_METRICS:
                if m not in applicable:
                    cur = latest.get(m)
                    if cur is None or cur.status == "MISSING":
                        save_metric(run_id, ch.id, d, m, None, "collector", status="NOT_AVAILABLE")
            if miss:
                missing_by_channel[ch.name] = miss
        else:
            # у канала нет подключённого источника вовсе (не ошибка сбора)
            no_source.append(ch.name)
            for m in DAILY_METRICS:
                cur = latest.get(m)
                if cur is not None and cur.status == "NOT_AVAILABLE":
                    continue
                save_metric(run_id, ch.id, d, m, None, "collector", status="NOT_AVAILABLE")
    db.session.commit()
    if no_source:
        exists = Notification.query.filter(
            Notification.message.like(f"%источник не подключён за {d}%")).first()
        if not exists:
            db.session.add(Notification(level="warn", message=(
                f"⚠ По каналам без источника данных за {d}: {', '.join(no_source)} — "
                "подключите коннектор или вносите CSV/вручную (экраны «Импорт»/«Ручной ввод»).")))
            db.session.commit()
    if missing_by_channel:
        exists = Notification.query.filter(
            Notification.message.like(f"%не получены данные за {d}%")).first()
        if not exists:
            channels_list = ", ".join(sorted(missing_by_channel))
            db.session.add(Notification(level="error", message=(
                f"❌ Не получены данные за {d} по каналам: {channels_list} "
                f"(всего {sum(len(v) for v in missing_by_channel.values())} метрик MISSING — "
                "импорт CSV или ручной ввод на соответствующих экранах).")))
            db.session.commit()
    return f"MISSING: {sum(len(v) for v in missing_by_channel.values())} метрик по {len(missing_by_channel)} каналам"


def daily_digest():
    """Короткая сводка за вчера — в очередь уведомлений (доставка в Telegram)."""
    from datetime import timedelta as _td
    from db import Registration, GcPayment
    d = date.today() - _td(days=1)
    if Notification.query.filter(Notification.message.like(f"Сводка за {d}%")).first():
        return
    from calc import aggregate, registrations_total
    agg, _ = aggregate(d, d)
    regs = registrations_total(d, d)
    pays = db.session.query(db.func.coalesce(db.func.sum(GcPayment.amount), 0)).filter(
        GcPayment.date == d, GcPayment.status == "accepted").scalar()
    db.session.add(Notification(level="info", message=(
        f"Сводка за {d}: охват {agg.get('reach') or 0:,.0f}, регистрации {regs:,.0f}, "
        f"оплаты {pays:,.0f} ₽.".replace(",", " "))))
    db.session.commit()


def run_daily_collection():
    run_id = start_run("daily_collect")
    results = [collect_youtube(run_id), mark_missing(run_id)]
    try:
        daily_digest()
    except Exception:
        pass
    finish_run(run_id, "; ".join(results))
    return results


# --------- XLSX-импорт (ПЛАН-ОТЧЁТ. Статистики СММ) ---------
import datetime as _dt
import re as _re

XLSX_CHANNEL_RULES = [
    ("alexey_syrover", "Алексей старый"), ("alexeynew", "Алексей новый"),
    ("alexeysyrover", "Алексей новый"), ("syrover_school", "Дина"), ("дина", "Дина"),
    ("венера", "ВЕНЕРА"), ("вконтакте", "VK"), ("телеграм", "Telegram"),
    ("youtube", "YouTube"), ("tiktok", "TikTok"), ("тик", "TikTok"), ("дзен", "Дзен"),
    ("max", "MAX основной"),
]
XLSX_SOURCE_RULES = {
    "insta-alexey": "insta-alexey", "insta-alexeynew": "insta-alexeynew",
    "insta-dina": "insta-dina", "insta-venera": "insta-venera",
    "telegram": "telegram", "vkontakte": "vkontakte", "vk": "vkontakte",
    "ticktock": "ticktock", "tiktok": "ticktock", "dzen": "dzen", "max": "max",
}
XLSX_CAMPAIGN_RULES = {"ОТКРЫТЫЙ УРОК": "openlesson", "ОТКРЫТЫЙ": "openlesson",
                       "БЕСПЛАТН": "openlesson", "ОТВЕТЫ": "otvety",
                       "ВОПРОС": "otvety", "АВТОВЕБ": "autoweb", "АВТО": "autoweb"}
RU_MONTHS = {"январ": 1, "феврал": 2, "март": 3, "апрел": 4, "мая": 5, "ма": 5, "июн": 6,
             "июл": 7, "август": 8, "сентябр": 9, "октябр": 10, "ноябр": 11, "декабр": 12}


def _xlsx_channel(name):
    low = (name or "").lower()
    for rule, target in XLSX_CHANNEL_RULES:
        if rule in low:
            return Channel.query.filter(Channel.name == target).first()
    return None


def _xlsx_date(v, default_year):
    if isinstance(v, _dt.datetime):
        return v.date()
    if isinstance(v, _dt.date):
        return v
    s = str(v or "").strip().lower()
    m = _re.match(r"(\d{1,2})", s)
    mon = next((n for k, n in RU_MONTHS.items() if k in s), None)
    if m and mon:
        try:
            return _dt.date(default_year, mon, min(int(m.group(1)), 28))
        except ValueError:
            return None
    try:
        return _dt.date.fromisoformat(s[:10])
    except ValueError:
        return None


def import_xlsx(fileobj, default_year=None):
    """Импорт книги «ПЛАН-ОТЧЁТ. Статистики СММ»:
    - «Итоги недели*» / «Охваты месяц»: средние охваты каналов -> reach;
    - «Лиды*»: сетка источник x medium x кампания -> регистрации с UTM;
    - «ТГ подписчики» / «вк»: дата+число -> followers;
    - универсальный лист date,channel_id,metric,value."""
    import openpyxl
    default_year = default_year or _dt.date.today().year
    wb = openpyxl.load_workbook(fileobj, data_only=True, read_only=True)
    run_id = start_run("xlsx_import")
    stats = {"reach": 0, "leads": 0, "followers": 0, "generic": 0, "skipped": []}
    def sheet_key(n):
        """Порядок обработки: охватные листы по возрастанию года (2023 → 2024 → 2526),
        чтобы свежие данные учебного года 25/26 перезаписывали старые на тех же датах."""
        low = n.lower()
        is_reach = low.startswith("итоги") or "охваты месяц" in low
        if "2023" in low:
            y = 0
        elif "2024" in low:
            y = 1
        elif is_reach:
            y = 2
        else:
            y = 3
        return (0 if is_reach else 1, y)

    for name in sorted(wb.sheetnames, key=sheet_key):
        ws = wb[name]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        low = name.lower()
        try:
            if low.startswith("итоги") or "охваты месяц" in low:
                stats["reach"] += _xlsx_parse_reach(rows, run_id, default_year)
            elif low.startswith("лиды"):
                stats["leads"] += _xlsx_parse_leads(rows, run_id, default_year)
            elif "подписчики" in low or low == "вк":
                stats["followers"] += _xlsx_parse_followers(rows, run_id, name)
            else:
                stats["generic"] += _xlsx_parse_generic(rows, run_id)
        except Exception as e:
            stats["skipped"].append(f"{name}: {str(e)[:60]}")
    finish_run(run_id, str(stats))
    return stats


def _xlsx_parse_reach(rows, run_id, default_year):
    n = 0
    if not rows:
        return 0
    header = rows[0]
    cols = []
    for idx, h in enumerate(header):
        if not h:
            continue
        txt = str(h)
        if "рост" in txt.lower() or "средние охват" in txt.lower():
            continue
        ch = _xlsx_channel(txt)
        if ch:
            cols.append((idx, ch))
    for row in rows[1:]:
        if not row:
            continue
        d = None
        for v in row[:3]:
            d = _xlsx_date(v, default_year)
            if d:
                break
        if not d:
            continue
        for idx, ch in cols:
            if idx < len(row) and isinstance(row[idx], (int, float)) and row[idx] > 0:
                save_metric(run_id, ch.id, d, "reach", float(row[idx]), "xlsx")
                n += 1
    db.session.commit()
    return n


def _xlsx_parse_leads(rows, run_id, default_year):
    from db import Registration
    n = 0
    if len(rows) < 5:
        return 0
    camps = {}
    cur = None
    for idx, v in enumerate(rows[0]):
        if v:
            txt = str(v).upper()
            m = next((c for k, c in XLSX_CAMPAIGN_RULES.items() if k in txt), None)
            if m:
                cur = m
        if cur:
            camps[idx] = cur
    med_row = None
    for r in rows[:5]:
        vals = [str(v or "").lower() for v in r]
        if any(x in vals for x in ("post", "canal-anons", "taplink", "gayd")):
            med_row = vals
            break
    if not med_row:
        return 0
    today = _dt.date.today()
    cur_date = today
    for row in rows[1:]:
        if not row:
            continue
        # строка-разделитель с датой недели ("7-14 апреля") — запоминаем
        first = str(row[0] or "").strip()
        if _re.match(r"\d{1,2}\s*[-–—]\s*\d{1,2}", first):
            d = _xlsx_date(first.split("–")[0].split("-")[0].strip() + " " +
                           " ".join(first.split()[1:]), default_year)
            if d:
                cur_date = d
            continue
        src_raw = first.lower()
        src = XLSX_SOURCE_RULES.get(src_raw) or XLSX_SOURCE_RULES.get(src_raw.replace(" ", ""))
        if not src:
            for k, v in XLSX_SOURCE_RULES.items():
                if k in src_raw:
                    src = v
                    break
        if not src:
            continue
        for idx, v in enumerate(row):
            if idx == 0 or not isinstance(v, (int, float)) or v <= 0 or idx >= len(med_row):
                continue
            medium = (med_row[idx] or "").strip()
            if medium in ("", "nan", "none", "общее"):
                continue
            db.session.add(Registration(date=cur_date, utm_source=src, utm_medium=medium,
                                        utm_campaign=camps.get(idx, ""), count=float(v),
                                        status="OK", landing="xlsx-import"))
            n += 1
    db.session.commit()
    return n


def _xlsx_parse_followers(rows, run_id, sheet_name):
    low = sheet_name.lower()
    if "тг" in low or "телеграм" in low:
        ch = Channel.query.filter(Channel.name == "Telegram").first()
    elif low == "вк":
        ch = Channel.query.filter(Channel.name == "VK").first()
    else:
        ch = None
    if not ch:
        return 0
    n = 0
    year = _dt.date.today().year
    for row in rows:
        if not row or len(row) < 2:
            continue
        d = _xlsx_date(row[0], year)
        if d and isinstance(row[1], (int, float)) and row[1] > 0:
            save_metric(run_id, ch.id, d, "followers", float(row[1]), "xlsx")
            n += 1
    db.session.commit()
    return n


def _xlsx_parse_generic(rows, run_id):
    from db import Registration
    if not rows:
        return 0
    head = [str(x or "").strip().lower() for x in rows[0]]
    if "metric" not in head or "value" not in head:
        return 0
    ci = {k: head.index(k) for k in ("date", "channel_id", "metric", "value") if k in head}
    n = 0
    for row in rows[1:]:
        try:
            d = _xlsx_date(row[ci["date"]], _dt.date.today().year)
            if not d:
                continue
            m = str(row[ci["metric"]]).strip()
            if m not in DAILY_METRICS:
                continue
            v = row[ci["value"]]
            save_metric(run_id, int(row[ci["channel_id"]]), d, m,
                        float(v) if v not in (None, "") else None, "xlsx")
            n += 1
        except Exception:
            continue
    db.session.commit()
    return n


ALLOWED_METRICS = set(DAILY_METRICS)


def import_csv_channel(fileobj):
    """CSV-импорт дневной статистики. Формат header:
    date,channel_id,metric,value[,status]
    date,registrations,utm_source,utm_medium,utm_campaign,landing,count  — для регистраций
    """
    run_id = start_run("csv_import")
    text = fileobj.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    n = 0
    for row in reader:
        row = {k.strip().lower(): (v.strip() if isinstance(v, str) else v) for k, v in row.items() if k}
        if not row.get("date"):
            continue
        d = datetime.strptime(row["date"][:10], "%Y-%m-%d").date()
        if row.get("metric"):
            if row["metric"] not in ALLOWED_METRICS:
                continue
            value = float(row["value"]) if row.get("value") not in (None, "",) else None
            status = row.get("status", "MANUAL" if value is not None else "MISSING").upper()
            save_metric(run_id, int(row["channel_id"]), d, row["metric"], value, "csv", status)
            n += 1
        elif row.get("registrations") or row.get("count"):
            from db import Registration
            cnt = float(row.get("count") or row.get("registrations") or 0)
            db.session.add(Registration(date=d, utm_source=row.get("utm_source", ""),
                                        utm_medium=row.get("utm_medium", ""),
                                        utm_campaign=row.get("utm_campaign", ""),
                                        landing=row.get("landing", ""), count=cnt, status="OK"))
            n += 1
    db.session.commit()
    finish_run(run_id, f"импортировано {n} строк")
    return n
